import openpyxl
import matplotlib.pyplot as plt
import numpy as np

# 读取Excel文件
wb = openpyxl.load_workbook("/home/root1/wcc/Langchain-Chatchat/save_excel/test.xlsx")
ws = wb.active

# 获取列名
column_names = ws[1].column_names

# 获取数据
data = []
for row in ws.iter_rows(min_row=2, values_only=True):
    data.append(row)

# 提取x和y列的数据
x_data = [row[0] for row in data]
y_data = [row[1] for row in data]

# 创建折线图
plt.plot(x_data, y_data)
plt.xlabel("x")
plt.ylabel("y")
plt.title("Line Plot")
plt.savefig("/home/root1/wcc/Langchain-Chatchat/coding/test.png")

# 显示图形
plt.show()
